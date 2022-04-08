import sys,os
# import gzip
import argparse
# from types import WrapperDescriptorType
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import cell, get_column_letter
# import logging
import subprocess
import numpy as np

tabix="tabix"
python3 = sys.executable
normaldepth=os.path.split(os.path.realpath(__file__))[0] + "/normalization_depth_from_mosdepth_output.py"
# part_len="/public/agis/chengshifeng_group/fengcong/WGRS/software/Fc-code/partlen.txt"

def depthp2color(depth):
    #根据深度返回颜色
    #深度范围0-2，超过2为红色
    color_list = ['BEBEBE','C4C4AA','CBCB96','D2D282','D9D96E','E0E05A','E7E746','EDED32','F4F41E','FBFB0A','FFF100',
        'FFD600','FFBB00','FFA100','FF8600','FF6B00','FF5000','FF3500','FF1A00','FF0000']
    if depth >= 2:
        return color_list[-1]
    else:
        step = 2.0/len(color_list)
        return color_list[int(depth/step)]


if __name__ == "__main__":
    cmdparser = argparse.ArgumentParser(description="RDV to excel/fengcong@caas.cn" )
    #input data
    cmdparser.add_argument("-r","--region", dest="region",type=str, required=True,
                            help="plot region")
    # cmdparser.add_argument("-g","--gff", dest="gff",type=str, required=False,
    #                         help="gff file if u want to plot gene structure")
    cmdparser.add_argument("-D","--Depth", dest="Depth",type=str, required=True,
                            help="Depth dir")
    ## No matter what, deep information must be drawn.
    cmdparser.add_argument("-d","--order", dest="order",type=str, required=True,
                            help="sample  order information ")
    #excel setting
    cmdparser.add_argument("-e","--excel", dest="excel",type=str, required=False,
                            help="input excel file.[optional]")
    cmdparser.add_argument("-c","--cellbp", dest="cellbp",type=int,default=1, required=False,
                                help="How many bp does a cell represent.default=1 [optional]")
    cmdparser.add_argument("-x","--cellrow", dest="cellrow",type=int,default=1, required=False,
                            help="How many rows in a cell.default=1 [optional]")
    cmdparser.add_argument("-s","--startrowcol", dest="startrowcol",type=str, default="1,1" , required=False,
                            help="start row and col.default=1,1 [optional]")
    
    #resouce
    cmdparser.add_argument("-t","--tabix", dest="tabix",type=str, required=False,default=tabix,
                            help="tabix path [optional]")

    cmdparser.add_argument("-o","--output", dest="output",type=str, required=True,
                            help="output xlsx file name")


    args = cmdparser.parse_args()

    #deal args
    ## region
    region = args.region
    chromosome = region.split(":")[0]
    start = int(region.split(":")[1].split("-")[0])
    end = int(region.split(":")[1].split("-")[1])

    need_region=[start, end]
    need_chr=chromosome

    # part_len_d={}
    # inf = open(part_len,"r")
    # for line in inf.readlines():
    #     part_len_d[line.strip().split()[0]] = int(line.strip().split()[1])
    # inf.close()
    # chro = chromosome
    # if chro!="chrUn":
    #     ## didnt check the region whether cross the part chr ##
    #     ## this can be a bug ##
    #     ## if u find that this isnt right, u should fix ths bug ##
    #     if need_region[0]  > part_len_d[chro+"_part1"]:
    #         need_chr = chro+"_part2"
    #         need_region[0] = need_region[0]-part_len_d[chro+"_part1"]
    #         need_region[1] = need_region[1]-part_len_d[chro+"_part1"]
    #     else:
    #         need_chr = chro+"_part1"
    # else:
    #     pass

    region="{0}:{1}-{2}".format(need_chr, need_region[0], need_region[1])

    ## sample order
    # read sample order file
    sample_order = []
    with open(args.order) as f:
        for line in f:
            sample_order.append(line.strip())



    

    ## get and normalize depth
    sample_depth={}
    header_line = []
    for sample in sample_order:
        header_line = []
        sample_depth[sample]=[]
        # print("""tail -n 1 %s | cut -f 4 """ %(args.Depth+"/"+sample+"/%s.mosdepth.summary.txt"%(sample)))
        child1 = subprocess.Popen(
        # """tail -n 1 %s | cut -f 4 """ %(args.Depth+"/"+sample+"/%s.Q0.mosdepth.summary.txt"%(sample)),shell=True,stdout=subprocess.PIPE
        """tail -n 1 %s | cut -f 4 """ %(args.Depth+"/"+sample+ "/"+[n for n in os.listdir(args.Depth+"/"+sample) if n.endswith("mosdepth.summary.txt")][0]),shell=True,stdout=subprocess.PIPE
        
        )
        averge_depth = child1.communicate()[0].decode("utf-8").strip()

        # print("""%s %s %s |
        #     %s %s -r %s -a %s -b %d  -s 
        #     """%(tabix , args.Depth+"/"+sample+"/%s.Q0.per-base.bed.gz"%(sample) , region,
        #         python3 , normaldepth,region,averge_depth,args.cellbp)
        #     )
        # child2 = subprocess.Popen("""%s %s %s |
        #     %s %s -r %s -a %s -b %d  -s 
        #     """%(tabix , args.Depth+"/"+sample+"/%s.Q0.per-base.bed.gz"%(sample) , region,
        #         python3 , normaldepth,region,averge_depth,args.cellbp),
        #     shell=True,stdout=subprocess.PIPE)
        
        child2 = subprocess.Popen("""%s %s %s |
            %s %s -r %s -a %s -b %d  -s 
            """%(tabix , args.Depth+"/"+sample+"/" +[n for n in os.listdir(args.Depth+"/"+sample) if n.endswith("per-base.bed.gz")][0] , region,
                python3 , normaldepth,region,averge_depth,args.cellbp),
            shell=True,stdout=subprocess.PIPE)

        for line in child2.stdout.readlines():
            depth = line.decode("utf-8").strip().split()[1]
            start_pos = line.decode("utf-8").strip().split()[0]
            start_pos = int(start_pos)
            depth = round(float(depth),2)
            header_line.append(start_pos)
            sample_depth[sample].append(depth)
    
    #get excel
    if args.excel:
        wb = openpyxl.load_workbook(args.excel)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active

    #start col and row
    header_row=int(args.startrowcol.split(",")[0])
    start_row = header_row+1
    start_col = int(args.startrowcol.split(",")[1])
    cell_row = args.cellrow

    #write header
    algn_center_rot = Alignment(horizontal="center",vertical="bottom",wrap_text=False,textRotation=90)
    for  index,col in enumerate( range(start_col,start_col+len(sample_depth[sample])) ):
        ws.cell(row=header_row,column=col).value = header_line[index]
        ws.cell(row=header_row,column=col).alignment = algn_center_rot
        ws.cell(row=header_row,column=col).comment = Comment("%d"%(header_line[index]), "fengcong@caas.cn")


    # first_row = True
    for index1,sample in enumerate(sample_order):
        # print(index1,sample)
        for index2, col in enumerate( range(start_col,start_col+len(sample_depth[sample])) ):
            # if first_row:
            #     first_row = False
            #     col_let = get_column_letter(col)
            #     ws.column_dimensions[col_let].width = 0.5
            
            for i in range(cell_row):
                col_let = get_column_letter(col)
                ws.column_dimensions[col_let].width = 0.5

                ws.cell(row=start_row+index1*cell_row+i,column=col).value = sample_depth[sample][index2]
                # ws.cell(row=start_row+index1*cell_row+i,column=col).comment=Comment("sample:%s\nbin_start:%d\naverage_depth:%.2f"%(sample,header_line[index2],round(np.mean(sample_depth[sample]),2)), 
                #     "fengcong@caas.cn", width=160, height=70)
                color_depth = depthp2color(sample_depth[sample][index2])
                color_filldp = PatternFill(fgColor=color_depth, fill_type="solid")
                ws.cell(row=start_row+index1*cell_row+i,column=col).fill = color_filldp
                # print(start_row+index1*i,col)
        ws.cell(row=start_row+index1*cell_row+i,column=col+1).value = round(np.mean(sample_depth[sample]),2)
        ws.cell(row=start_row+index1*cell_row+i,column=col+1).comment = Comment("sample:%s\naverage_depth:%.2f"%(sample,round(np.mean(sample_depth[sample]),2)), 
            "fengcong@caas.cn",width=160, height=70)

    wb.save(args.output)
        
            

