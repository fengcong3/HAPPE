#2021-8-3 15:21:47
# fengcong@caas.cn
# write sample information to excel
# speed up while cellrow > 1

import sys
import argparse
# from types import WrapperDescriptorType
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill
from openpyxl.utils import cell, get_column_letter

if __name__ == "__main__":
    cmdparser = argparse.ArgumentParser(description="write sample information to excel/fengcong@caas.cn" )
    cmdparser.add_argument("-o","--output", dest="output",type=str, required=True,
                            help="output xlsx file name")
    cmdparser.add_argument("-e","--excel", dest="excel",type=str, required=False,
                            help="input excel file.[optional]")
    cmdparser.add_argument("-i","--info", dest="information",type=str, required=True,
                            help="individual information,sample number/name must constant with the tree file.")
    cmdparser.add_argument("-r","--cellrow", dest="cellrow",type=int,default=1, required=False,
                            help="How many rows in a cell.default=1 [optional]")
    cmdparser.add_argument("-s","--startrowcol", dest="startrowcol",type=str, default="1,1" , required=False,
                            help="How many rows in a cell.default=1,1 [optional]")
    cmdparser.add_argument("-d","--order", dest="order",type=str, required=True,
                            help="sample order ")
    cmdparser.add_argument("-c","--color", dest="color",type=str, required=False,
                            help="individual color,sample number/name must constant with the tree file.[optional]")

    args = cmdparser.parse_args()
    # deal args
    if args.excel:
        wb = openpyxl.load_workbook(args.excel)
    else:
        wb = openpyxl.Workbook()
    
    start_row = int(args.startrowcol.split(",")[0])
    start_col = int(args.startrowcol.split(",")[1])
    cell_row = args.cellrow

    color_d = {}
    if args.color:
        inf = open(args.color,"r")
        for line in inf.readlines():
            ls = line.strip().split()
            color_d[ls[0]] = ls[1]
        inf.close()

    
    #open a sheet
    ws = wb.active
    # print(ws.title)
    # print(ws.max_row)
    # print(ws.max_column)

    # read sample order file
    sample_order = []
    with open(args.order) as f:
        for line in f:
            sample_order.append(line.strip())

    # read sample information
    inf = open(args.information,"r")
    sample_info = {}

    header_line = inf.readline().strip().split("\t")
    for line in inf:
        line = line.strip().split("\t")
        sample_info[line[0]] = line[1:]
    
    inf.close()

    algn_center = Alignment(horizontal="center",vertical="center",wrap_text=False)
    light_ft = Font(name='Times New Roman', size=12 , bold=False)
    bold_ft = Font(name='Times New Roman', size=12 , bold=True)
    

    #write header
    for index,i in enumerate(header_line):
        ws.cell(row=start_row,column=start_col+index).value = i
        ws.cell(row=start_row,column=start_col+index).font = bold_ft
        ws.cell(row=start_row,column=start_col+index).alignment = algn_center
    sys.stdout.write("%d\n"%(start_col+index)) # if u want to add something after this , the start col show be (start_col+ index +1)
    #write sample information
    first_row = True
    for index,sample in enumerate(sample_order):
        merge_range_item = [cell_row * index + start_row+ 1, start_col,cell_row * (index+1)-1  + start_row+ 1,start_col] 
        if cell_row >1 and first_row and False:
            ws.merge_cells(start_row=merge_range_item[0], start_column=merge_range_item[1], 
                end_row=merge_range_item[2], end_column=merge_range_item[3])
            ws.cell(row=merge_range_item[0],column=merge_range_item[1]).value = sample
            ws.cell(row=merge_range_item[0],column=merge_range_item[1]).font = bold_ft
            ws.cell(row=merge_range_item[0],column=merge_range_item[1]).alignment = algn_center
            col_let = get_column_letter(merge_range_item[1])
            ws.column_dimensions[col_let].width = 14.25
            if len(color_d) != 0:
                color_fill = PatternFill(fgColor=color_d[sample], fill_type="solid")
                # print(sample,color_d[sample])
                ws.cell(merge_range_item[0],merge_range_item[1]).fill = color_fill
        else:
            for i in range(cell_row):
                ws.cell(row=merge_range_item[0]+i,column=merge_range_item[1]).value = sample
                ws.cell(row=merge_range_item[0]+i,column=merge_range_item[1]).font = bold_ft
                ws.cell(row=merge_range_item[0]+i,column=merge_range_item[1]).alignment = algn_center
                col_let = get_column_letter(merge_range_item[1])
                ws.column_dimensions[col_let].width = 14.25
                if len(color_d) != 0 and sample in color_d:
                    color_fill = PatternFill(fgColor=color_d[sample], fill_type="solid")
                    # print(sample,color_d[sample])
                    ws.cell(merge_range_item[0]+i,merge_range_item[1]).fill = color_fill
        if sample in sample_info:
            info_list=sample_info[sample]
        else:
            info_list = [""] * len(header_line)

        for index1,value in enumerate(info_list):
            merge_range_item1 = [cell_row * index + start_row+ 1, start_col+index1+1, cell_row * (index+1)-1  + start_row+ 1,start_col+index1+1] 
            if cell_row >1 and first_row and False:
                ws.merge_cells(start_row=merge_range_item1[0], start_column=merge_range_item1[1],
                    end_row=merge_range_item1[2], end_column=merge_range_item1[3])
                ws.cell(row=merge_range_item1[0],column=merge_range_item1[1]).value = value
                ws.cell(row=merge_range_item1[0],column=merge_range_item1[1]).font = light_ft
                ws.cell(row=merge_range_item1[0],column=merge_range_item1[1]).alignment = algn_center
            else:
                for i in range(cell_row):
                    ws.cell(row=merge_range_item1[0]+i,column=merge_range_item1[1]).value = value
                    ws.cell(row=merge_range_item1[0]+i,column=merge_range_item1[1]).font = light_ft
                    ws.cell(row=merge_range_item1[0]+i,column=merge_range_item1[1]).alignment = algn_center
        first_row = False
    #save file
    wb.save(args.output)

