#2021-6-18 16:36:54
#fengcong@caas.cn
#plot tree in excel,newick tree ,could have bootstrap


# -c --color
# sample1    990099
# sample2    009900
# import re
import sys,os
import argparse
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill
from openpyxl.utils import   get_column_letter

sys.setrecursionlimit(100000)
def find_comma_index(tree_seq):
    comma_index_list = []
    for index,i in enumerate(tree_seq):
        if i == ",":
            comma_index_list.append(index)

    i_list = []
    for i in comma_index_list:
        if tree_seq[0:i].count("(") == tree_seq[0:i].count(")")  and tree_seq[i+1:].count(")") == tree_seq[i+1:].count("(") :
            # print(111)
            i_list.append(i)

    return i_list
        

# def rm_bracket(tree_seq):
#     if tree_seq[0] == "(":
#         rbound = tree_seq.rfind(")")
#         return tree_seq[1:rbound]
#     else:
#         return tree_seq

def rm_bracket2(tree_seq):
    if tree_seq[0] == "(":
        rbound = tree_seq.rfind(")")
        if rbound == len(tree_seq)-1:
            return (tree_seq[1:-1],1)
        else:
            return (tree_seq[1:rbound],float(tree_seq[rbound+1:-1].split(":")[-1]))
    else:
        if ":" in tree_seq:
            return (tree_seq,float(tree_seq.split(":")[-1]))
        else:
            return (tree_seq,1)


def post_traverse_tree(tree_seq,cur_branch_len):
    comma_index = find_comma_index(tree_seq)
    comma_index_len = len(comma_index)
    # print(tree_seq,comma_index)

    if comma_index_len == 0:
        sample_name = tree_seq.split(":")[0]
        return ([sample_name],cur_branch_len)


    # split the child node
    seq_list = []
    pre_start = 0
    for i,indx in enumerate(comma_index) :
        (sub_tree_seq,sub_branch_len) = rm_bracket2( tree_seq[pre_start:indx] )
        seq_list.append( (sub_tree_seq,sub_branch_len)  )
        pre_start = indx+1
        if i == len(comma_index) - 1:
            sub_tree_seq,sub_branch_len = rm_bracket2( tree_seq[pre_start:] )
            seq_list.append( (sub_tree_seq,sub_branch_len)  )
    
    sample_list = []
    cur_max_branch_len = 0
    for i,(sub_tree_seq,sub_branch_len) in enumerate(seq_list):
        sub_sample_list , sub_max_depth = post_traverse_tree(sub_tree_seq,cur_branch_len + sub_branch_len)
        sample_list.extend(sub_sample_list)
        cur_max_branch_len = max(cur_max_branch_len,sub_max_depth)
    
    return (sample_list,cur_max_branch_len)

#define the border style
border_bot = Border(bottom=Side(border_style="medium",color="000000"))
border_rig = Border(right=Side(border_style="medium",color="000000"))
border_rig_bot = Border(right=Side(border_style="medium",color="000000"),bottom=Side(border_style="medium",color="000000"))
border_lef_bot = Border(left=Side(border_style="medium",color="000000"),bottom=Side(border_style="medium",color="000000"))
border_lef_bot_rig = Border(left=Side(border_style="medium",color="000000"),bottom=Side(border_style="medium",color="000000"), \
    right=Side(border_style="medium",color="000000"))  
border_bot_dash = Border(bottom=Side(border_style="mediumDashed",color="D3D3D3"))
tree_offset = 1
tree_width = 1000

def post_traverse_tree_and_plot(tree_seq,samples_up_to_you,father_depth,cur_depth,cell_unit,working_sheet):
    comma_index = find_comma_index(tree_seq)
    comma_index_len = len(comma_index)
    # print(tree_seq,comma_index)
    if comma_index_len == 0: #leaf node
        sample_name = tree_seq.split(":")[0]
        samples_up_to_you.append(sample_name)
        row_num = 2*len(samples_up_to_you) -1
        father_col = int(father_depth/(cell_unit*1.0))
        cur_col = int(cur_depth/(cell_unit*1.0))
        col_num = [x for x in range(father_col+1,cur_col+1)]
        # print(sample_name,father_depth,cur_depth,father_col,cur_col,col_num) #########
        for i in col_num:
            working_sheet.cell(row=row_num,column=i+tree_offset).border = border_bot
        for k in range(cur_col+1,tree_width):
            working_sheet.cell(row=row_num,column=k+tree_offset).border = border_bot_dash

        return (samples_up_to_you,row_num)
    
    # split the child node
    seq_list = []
    pre_start = 0
    for i,indx in enumerate(comma_index) :
        (sub_tree_seq,sub_branch_len) = rm_bracket2( tree_seq[pre_start:indx] )
        seq_list.append( (sub_tree_seq,sub_branch_len)  )
        pre_start = indx+1
        if i == len(comma_index) - 1:
            sub_tree_seq,sub_branch_len = rm_bracket2( tree_seq[pre_start:] )
            seq_list.append( (sub_tree_seq,sub_branch_len)  )

    #traverse the child node
    row_num_list = []
    for i,(sub_tree_seq,sub_branch_len) in enumerate(seq_list):
        samples_up_to_you , row_num = post_traverse_tree_and_plot(sub_tree_seq,samples_up_to_you,cur_depth,sub_branch_len+cur_depth,cell_unit,working_sheet)
        row_num_list.append(row_num)
    
    #plot the branch and vertical line
    father_col = int(father_depth/(cell_unit*1.0))
    cur_col = int(cur_depth/(cell_unit*1.0))
    col_num_list = [x for x in range(father_col+1,cur_col+1)]
    row_num = int( (row_num_list[0]+row_num_list[-1])/2 )
    # print("internode",father_depth,cur_depth,father_col,cur_col,col_num_list,row_num) #########
    ## plot vertical line
    
    
    for row in range(row_num_list[0]+1,row_num_list[-1]+1):
        if row == row_num and cur_col != 1 and col_num_list != []:
            working_sheet.cell(row,cur_col+tree_offset).border = border_rig_bot
        else:
            working_sheet.cell(row,cur_col+tree_offset).border = border_rig
    ## plot branch
    for index in range(len(col_num_list)-1):
        working_sheet.cell(row_num,col_num_list[index]+tree_offset).border = border_bot
    
    if father_depth == 0 and cur_depth == 0:
        # print(row_num_list[0],row_num_list[-1])
        for row in range(row_num_list[0]+1,row_num_list[-1]+1):
            working_sheet.cell(row,0+tree_offset).border = border_rig

    return (samples_up_to_you,row_num)
    
                                                                                                                                                                                                         



if __name__ == "__main__":
    cmdparser = argparse.ArgumentParser(description="plot tree in excel table/fengcong@caas.cn" )
    cmdparser.add_argument("-o","--output", dest="output",type=str, required=True,
                            help="output xlsx file name")
    cmdparser.add_argument("-c","--color", dest="color",type=str, required=False,
                            help="individual color,sample number/name must constant with the tree file.[optional]")
    cmdparser.add_argument("-i","--insert", dest="insert",type=int, default=2, required=False,
                            help="insert N rows at header.(default=2)[optional]")
    cmdparser.add_argument("-w","--width", dest="width",type=int, default=1000, required=False,
                            help="How many columns do you want to occupy for this tree topology.(default=1000)[optional]")
    # cmdparser.add_argument("-i","--info", dest="information",type=str, required=False,
    #                         help="individual information,sample number/name must constant with the tree file.[optional]")
    cmdparser.add_argument('trees', metavar='trees', type=str, nargs='+',
                    help='tree files,newick format')

    args = cmdparser.parse_args()

    #deal args
    tree_files_list = args.trees
    output_file_name = args.output
    color_file = args.color
    tree_width = args.width
    # tree_files_list = ["/vol2/agis/chengshifeng_group/fengcong/kmer_read_self_test/cYLeLh34Fj8fqALXY43Aaw_newick.txt"]
    # output_file_name= "/vol2/agis/chengshifeng_group/fengcong/kmer_read_self_test/169pure_add_223modern_missense.newick.xlsx"
    # color_file = "/vol2/agis/chengshifeng_group/fengcong/kmer_read_self_test/1059.color"
    # color_file=None

    #open workbook
    tree_workbook = openpyxl.Workbook()

    #recording sheet position
    sheet_pos = 0

    #calc sample count and get sample color 
    
    color_d = {}
    if color_file:
        inf = open(color_file,"r")
        for line in inf.readlines():
            ls = line.strip().split()
            color_d[ls[0]] = ls[1]
        inf.close()


    #traverse each tree file to different Sheet
    for tree_file in tree_files_list: 
        ##open the tree file
        tree_file_handle = open(tree_file,"r")
        lines = tree_file_handle.readlines()
        tree_file_handle.close()
        if len(lines) != 1:
            sys.stderr.write("this file has more than one line.\n")
            exit(-1)

        tree_seq = lines[0].strip()
        tree_seq = tree_seq.strip(";")
        tree_seq = tree_seq[1:-1]  #skip "(" and ")"

        # col_let = get_column_letter(start_col+index+2)
        # ws.column_dimensions[col_let].width = 4

        sheet_name = os.path.basename(tree_file).split(".")[0]
        working_sheet = tree_workbook.create_sheet(sheet_name,sheet_pos)
        #set tree width
        for col in range(2,tree_width+2):
            col_let = get_column_letter(col)
            working_sheet.column_dimensions[col_let].width = 0.5

        #we need posttraverse the tree
        sample_list,max_depth= post_traverse_tree(tree_seq,0)
        tree_go_left_offset = 50
        cell_unit = (max_depth*1.0)/(tree_width-1-tree_go_left_offset)
        # print("cell_unit:",cell_unit)
        # print("sample_List:",sample_list)

        post_traverse_tree_and_plot(tree_seq,[],0,0,cell_unit,working_sheet)

        
        max_d = tree_width 
        sys.stdout.write("%d\n"%(max_d)) # if u want to add some information f0llowing this tree, the start col of the information should be maxd+2
        working_sheet.insert_rows(1,args.insert) #insert 2 rows
        algn_center = Alignment(horizontal="center",vertical="center",wrap_text=False)
        light_ft = Font(name='Times New Roman', size=12 , bold=False)
        col_let = get_column_letter(max_d+1)
        working_sheet.column_dimensions[col_let].width = 17.75
        # # print(sample_list)
        for index,sample in enumerate(sample_list):
            merge_range_item = [2*(index+1)-1 + args.insert,max_d+1,2*(index+1)+args.insert,max_d+1] #[0,0,0,0] #start_row=None, start_column=None, end_row=None, end_column=None
            # print(merge_range_item)
            working_sheet.merge_cells(start_row=merge_range_item[0], start_column=merge_range_item[1],end_row=merge_range_item[2], end_column=merge_range_item[3])
            
            sys.stdout.write(sample+"\n")
            if len(color_d) != 0 and sample in color_d:
                color_fill = PatternFill(fgColor=color_d[sample], fill_type="solid")
                # print(sample,color_d[sample])
                working_sheet.cell(merge_range_item[0],merge_range_item[1]).fill = color_fill
            working_sheet.cell(merge_range_item[0],merge_range_item[1]).value = sample
            working_sheet.cell(merge_range_item[0],merge_range_item[1]).alignment  = algn_center
            working_sheet.cell(merge_range_item[0],merge_range_item[1]).font = light_ft
        sheet_pos +=1

        
    
    tree_workbook.save(filename=output_file_name)
