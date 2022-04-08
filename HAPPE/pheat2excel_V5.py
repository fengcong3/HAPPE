#2021-8-3 15:21:47
# fengcong@caas.cn
# write pheat matrix to excel
# speed up when cellrow > 1

import sys,os
import gzip
import argparse
from types import WrapperDescriptorType
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill
from openpyxl.utils import cell, get_column_letter
import SNPANN_upgrade_3 
from openpyxl.comments import Comment

bcftools = "/public/agis/chengshifeng_group/fengcong/WGRS/software/bcftools1.9/bin/bcftools"

if __name__ == "__main__":
    cmdparser = argparse.ArgumentParser(description="write pheat matrix to excel/fengcong@caas.cn" )
    cmdparser.add_argument("-i","--inpheat", dest="inpheat",type=str, required=True,
                            help="input pheat file.")
    cmdparser.add_argument("-o","--output", dest="output",type=str, required=True,
                            help="output xlsx file name")
    cmdparser.add_argument("-v","--vcf", dest="vcf",type=str, required=True,
                            help="input vcf file ,must be gziped")
    cmdparser.add_argument("-e","--excel", dest="excel",type=str, required=False,
                            help="input excel file.[optional]")
    cmdparser.add_argument("-r","--cellrow", dest="cellrow",type=int,default=1, required=False,
                            help="How many rows in a cell.default=1 [optional]")
    cmdparser.add_argument("-s","--startrowcol", dest="startrowcol",type=str, default="1,1" , required=False,
                            help="start row and col.default=1,1 [optional]")
    cmdparser.add_argument("-t","--cluster", dest="cluster",type=str, required=False,
                            help="sample  cluster information ")
    cmdparser.add_argument("-d","--order", dest="order",type=str, required=True,
                            help="sample  order information ")
    cmdparser.add_argument("-c","--color", dest="color",type=str, required=False,
                            help="individual color,sample number/name must constant with the tree file.[optional]")
    cmdparser.add_argument("-m","--moreinf", dest="moreinf",type=str, required=False,
                            help="more information about snp.[optional]")
    cmdparser.add_argument("-R","--Ref", dest="Ref",type=str, required=False,
                            help="change reference and color system.[optional]")
    cmdparser.add_argument("-F","--FuncAnn", dest="FuncAnn",type=str, required=False,
                            help="functional annotation.")


    args = cmdparser.parse_args()
    # deal args
    if args.excel:
        wb = openpyxl.load_workbook(args.excel)
    else:
        wb = openpyxl.Workbook()
    
    start_row = int(args.startrowcol.split(",")[0])
    start_col = int(args.startrowcol.split(",")[1])
    cell_row = args.cellrow
    
    color_file = args.color
    color_d = {}
    if color_file:
        inf = open(color_file,"r")
        for line in inf.readlines():
            ls = line.strip().split()
            color_d[ls[0]] = ls[1]
        inf.close()

    #func ann
    if args.FuncAnn :
        funcann = args.FuncAnn
    else:
        funcann = None

    # read order and cluster file
    cluster_d ={}
    if args.cluster:
        order_inf = open(args.cluster,"r")
        header = order_inf.readline()
        for line in order_inf:
            ls = line.strip().split()
            cluster_d[ls[0]] = ls[1]
        order_inf.close()


    #read pheat file
    snp_list = []
    snp_ann = {} # snp1:[G,G/T,T,intron_variant]
    sample_gt = {} # sample1:[-9,0,9,...]

    pheat_inf = open(args.inpheat,"r")
    header_line = pheat_inf.readline().strip().split()
    for sample in header_line[3:]:
        sample_gt[sample] = []
    
    col_num = 0
    for line in pheat_inf:
        col_num +=1
        ls = line.strip().split()
        snp_list.append(ls[0])
        for index,gt in enumerate(ls[3:]):
            sample_gt[ header_line[index+3] ].append(gt) 

    pheat_inf.close()

    ## get snp annotation
    for snp in snp_list:
        snp_ann[snp] = ["","","","","",""]  # ref , alt , feature_id , annotation , c.HGVS , p.HGVS
    
    # region_str = "%s:%s-%s"%(snp_list[0].split("_")[0], snp_list[0].split("_")[1] , snp_list[-1].split("_")[1])
    # # print(region_str)
    # snps = os.popen('%s view -r %s %s'%(bcftools , region_str,args.vcf))

    vcf_file = gzip.open(args.vcf,"rt") if args.vcf.endswith(".gz") else open(args.vcf, "r")
    line = vcf_file.readline()
    while line:
        if not line.startswith("#"):

            ls = line.split()
            # print(ls[2])
            if ls[2] in snp_ann:
                # print(1122)
                snp_ann[ls[2]][0] = ls[3]
                # snp_ann[ls[2]][1] = ls[3]+"/"+ls[4]
                snp_ann[ls[2]][1] = ls[4]
                ress=SNPANN_upgrade_3.SNP_ANN_of_Gene_Structure(ls[7])
                feature_id , annotation , cHGVS , pHGVS =ress  if ress else ["","","",""]

                
                snp_ann[ls[2]][2]= feature_id
                snp_ann[ls[2]][3] = annotation
                snp_ann[ls[2]][4] =cHGVS
                snp_ann[ls[2]][5] = pHGVS
        
        line = vcf_file.readline()

    # read func ann
    if funcann:
        funcannd = {}
        funcannf = open(funcann,"r")
        header = funcannf.readline()
        line = funcannf.readline()
        while line:
            ls = line.strip().split("\t")
            # ls[0] = ls[0].replace("01G","02G")
            funcannd[ls[0]] = ls[4]

            line = funcannf.readline()
        funcannf.close()

    #insert funcann to snpann
    if funcann:
        for snp in snp_ann:
            if snp_ann[snp][2] in funcannd:
                snp_ann[snp].insert(3,funcannd[ snp_ann[snp][2] ])
            else:
                snp_ann[snp].insert(3,"")
    
    info_num = 0
    snp_info = {} # {snpid:{p-value: 1e-5 , addi: -5, ...}}
    snp_info_field = []
    if args.moreinf:
        with open(args.moreinf) as f:
            #read header
            header_list = f.readline().strip().split()
            snp_info_field = header_list[1:]
            for line in f:
                ls = line.strip().split()
                snp_info[ls[0]] = {}
                for index , info in enumerate( header_list[1:] ):
                    snp_info[ ls[0] ][ info ] = ls[index + 1]

                if info_num == 0:
                    info_num = len(ls) - 1
                else:
                    if info_num == len(ls) - 1:
                        pass
                    else:
                        sys.stderr.write("different col num of snp information file.\n")
                        exit(-1)
    
    
    snp_info_row_num = 8 + info_num if funcann else 7 + info_num 

    #output pheat end rowcol
    sys.stdout.write("%d,%d\n"%(snp_info_row_num,start_col+col_num+2))

    #open a sheet
    ws = wb.active
    # print(ws.title)


    # read sample order file
    sample_order = []
    with open(args.order) as f:
        for line in f:
            sample_order.append(line.strip())



    algn_center_rot = Alignment(horizontal="center",vertical="bottom",wrap_text=False,textRotation=90)
    algn_center = Alignment(horizontal="center",vertical="center",wrap_text=False)
    light_ft = Font(name='Times New Roman', size=12 , bold=False)
    bold_ft = Font(name='Times New Roman', size=12 , bold=True)
    

    #write header
    ws.row_dimensions[1].height = 99.75
    # print(snp_list)
    tmp_header_line = ["","SNP ID"]
    tmp_header_line.extend(snp_list)
    for index,i in enumerate(tmp_header_line):
        ws.cell(row=start_row,column=start_col+index).value = i
        ws.cell(row=start_row,column=start_col+index).font = bold_ft
        ws.cell(row=start_row,column=start_col+index).alignment = algn_center_rot

    #write ann
    vcf_ann_list = ["allele A","allele B","Feature ID" ,"Gene Function", "Annotation" , "HGVS-nucleotide changes" , "HGVS-amino acid changes"] if funcann \
        else ["allele A","allele B","Feature ID" , "Annotation" , "HGVS-nucleotide changes" , "HGVS-amino acid changes"]
    vcf_ann_height = [51.75 , 51.75 , 158.25 , 95.25 ,158.25 , 95.25 , 95.25] if funcann \
        else [51.75 , 51.75 , 158.25 , 158.25 , 95.25 , 95.25]
    for aindex , x in enumerate(vcf_ann_list):
        # ws.cell(row=start_row+1+aindex,column=start_col).value = "Sample"
        ws.cell(row=start_row+1+aindex,column=start_col+1).value = x
        ws.cell(row=start_row+1+aindex,column=start_col+1).alignment = algn_center_rot
        ws.row_dimensions[start_row+1+aindex].height = vcf_ann_height[aindex]
        for index,i in enumerate(snp_list):
            col_let = get_column_letter(start_col+index+2)
            ws.column_dimensions[col_let].width = 4
            ws.cell(row=start_row+1+aindex,column=start_col+index+2).value = snp_ann[i][aindex]
            ws.cell(row=start_row+1+aindex,column=start_col+index+2).font = bold_ft
            ws.cell(row=start_row+1+aindex,column=start_col+index+2).alignment = algn_center_rot
    
    #write moreinfo
    for aindex , x in enumerate(snp_info_field):
        # ws.cell(row=start_row+1+aindex,column=start_col).value = "Sample"
        ws.cell(row=start_row+1+len(vcf_ann_list)+aindex,column=start_col+1).value = x
        ws.cell(row=start_row+1+len(vcf_ann_list)+aindex,column=start_col+1).alignment = algn_center_rot
        ws.row_dimensions[start_row+1+len(vcf_ann_list)+aindex].height = 99.75
        if aindex == len(snp_info_field) -1:
            ws.cell(row=start_row+1+len(vcf_ann_list)+aindex,column=start_col).value = "Sample"
            ws.cell(row=start_row+1+len(vcf_ann_list)+aindex,column=start_col).font = bold_ft
            ws.cell(row=start_row+1+len(vcf_ann_list)+aindex,column=start_col).alignment = algn_center
        for index,i in enumerate(snp_list):
            # col_let = get_column_letter(start_col+index+2)
            # ws.column_dimensions[col_let].width = 4
            ws.cell(row=start_row+1+len(vcf_ann_list)+aindex,column=start_col+index+2).value = snp_info[i][x] if i in snp_info else ""
            ws.cell(row=start_row+1+len(vcf_ann_list)+aindex,column=start_col+index+2).font = bold_ft
            ws.cell(row=start_row+1+len(vcf_ann_list)+aindex,column=start_col+index+2).alignment = algn_center_rot
    
    #write sample pheat
    first_row = True
    for index,sample in enumerate(sample_order):
        # print(sample)
        merge_range_item = [cell_row * index + start_row+ snp_info_row_num, start_col,cell_row * (index+1)-1 + start_row+ snp_info_row_num,start_col] 
        if cell_row > 1 and first_row and False:
            ws.merge_cells(start_row=merge_range_item[0], start_column=merge_range_item[1], 
                end_row=merge_range_item[2], end_column=merge_range_item[3])
            ws.cell(row=merge_range_item[0],column=merge_range_item[1]).value = sample
            ws.cell(row=merge_range_item[0],column=merge_range_item[1]).font = bold_ft
            ws.cell(row=merge_range_item[0],column=merge_range_item[1]).alignment = algn_center
            if len(color_d) != 0:
                color_fill = PatternFill(fgColor=color_d[sample], fill_type="solid")
                ws.cell(merge_range_item[0],merge_range_item[1]).fill = color_fill
        else:
            for i in range(cell_row):
                ws.cell(row=merge_range_item[0]+i,column=merge_range_item[1]).value = sample
                ws.cell(row=merge_range_item[0]+i,column=merge_range_item[1]).font = bold_ft
                ws.cell(row=merge_range_item[0]+i,column=merge_range_item[1]).alignment = algn_center
                if len(color_d) != 0 and sample in color_d:
                    color_fill = PatternFill(fgColor=color_d[sample], fill_type="solid")
                    ws.cell(merge_range_item[0]+i,merge_range_item[1]).fill = color_fill


        merge_range_item2 = [cell_row * index + start_row+ snp_info_row_num, start_col+1, cell_row * (index+1)-1  + start_row+ snp_info_row_num,start_col+1] 
        if cell_row > 1 and first_row and False:
            ws.merge_cells(start_row=merge_range_item2[0], start_column=merge_range_item2[1], 
                end_row=merge_range_item2[2], end_column=merge_range_item2[3])
            ws.cell(row=merge_range_item2[0],column=merge_range_item2[1]).value = cluster_d[sample] if cluster_d and sample in cluster_d else "NA"
            ws.cell(row=merge_range_item2[0],column=merge_range_item2[1]).font = bold_ft
            ws.cell(row=merge_range_item2[0],column=merge_range_item2[1]).alignment = algn_center
        else:
            for i in range(cell_row):
                ws.cell(row=merge_range_item2[0]+i,column=merge_range_item2[1]).value = cluster_d[sample] if cluster_d and sample in cluster_d else "NA"
                ws.cell(row=merge_range_item2[0]+i,column=merge_range_item2[1]).font = bold_ft
                ws.cell(row=merge_range_item2[0]+i,column=merge_range_item2[1]).alignment = algn_center
        
        # ref_index = sample_order.index(args.Ref) if args.Ref else None
        ref_gt_list = sample_gt[args.Ref] if args.Ref else ["-9" for n in range(len(snp_list))]
        for index1,value in enumerate(sample_gt[sample]):
            merge_range_item1 = [cell_row * index + start_row+ snp_info_row_num, start_col+index1+2, cell_row * (index+1)-1  + start_row+ snp_info_row_num,start_col+index1+2] 
            
            ref_gt = ref_gt_list[index1]
            reverse_gt = False
            colored = True
            if ref_gt == "-9":
                reverse_gt = False
            elif ref_gt == "0" or ref_gt =="NA":
                colored = False
            elif ref_gt == "9":
                reverse_gt = True
            else:
                sys.stderr.write("something wrong: %s , %s\n"%(__file__, sys._getframe().f_lineno))
                exit(-1)


            v = ""
            color_gt = "808080"
            if colored and not reverse_gt and value == "-9":
                v = snp_ann[ snp_list[index1] ][0]
                color_gt = "0000FF"
            elif colored and  reverse_gt and value == "-9":
                v = snp_ann[ snp_list[index1] ][0]
                color_gt = "FF0000"
            elif colored and value == "0":
                v = snp_ann[ snp_list[index1] ][0] + "/" + snp_ann[ snp_list[index1] ][1]
                color_gt = "FFFF00"
            elif colored and not reverse_gt and value == "9":
                v = snp_ann[ snp_list[index1] ][1]
                color_gt = "FF0000"
            elif colored and  reverse_gt and value == "9":
                v = snp_ann[ snp_list[index1] ][1]
                color_gt = "0000FF"
            elif not colored  and value == "9":
                v = snp_ann[ snp_list[index1] ][1]
                color_gt = "FFFFFF"
            elif not colored  and value == "-9":
                v = snp_ann[ snp_list[index1] ][0]
                color_gt = "FFFFFF"
            elif not colored  and value == "0":
                v = snp_ann[ snp_list[index1] ][0] + "/" + snp_ann[ snp_list[index1] ][1]
                color_gt = "FFFFFF"
            else:
                v = "NA"

            if cell_row > 1 and first_row and False:
                ws.merge_cells(start_row=merge_range_item1[0], start_column=merge_range_item1[1],
                    end_row=merge_range_item1[2], end_column=merge_range_item1[3])
            

                ws.cell(row=merge_range_item1[0],column=merge_range_item1[1]).value = v
                color_fillgt = PatternFill(fgColor=color_gt, fill_type="solid")

                ws.cell(merge_range_item1[0],merge_range_item1[1]).fill = color_fillgt
                ws.cell(row=merge_range_item1[0],column=merge_range_item1[1]).font = light_ft
                ws.cell(row=merge_range_item1[0],column=merge_range_item1[1]).alignment = algn_center
            else:
                for i in range(cell_row):
                    ws.cell(row=merge_range_item1[0]+i,column=merge_range_item1[1]).value = v
                    color_fillgt = PatternFill(fgColor=color_gt, fill_type="solid")

                    ws.cell(merge_range_item1[0]+i,merge_range_item1[1]).fill = color_fillgt
                    ws.cell(row=merge_range_item1[0]+i,column=merge_range_item1[1]).font = light_ft
                    ws.cell(row=merge_range_item1[0]+i,column=merge_range_item1[1]).alignment = algn_center

        first_row =False

    #save file
    wb.save(args.output)

